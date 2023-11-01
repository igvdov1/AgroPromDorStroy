import pandas as pd
from openpyxl import load_workbook, Workbook
import os 

def create_names_from_excel_files():
    works = {'АОСР', 'АРОО', 'АООК'}
    names_aosr = set()
    names_aook = set()
    names_aroo = set()
    for name in os.listdir('estimates'):
        xls = load_workbook(fr'estimates\{name}')
        ws = xls.active
        for i in range(1, len(ws['A']) + 1):
            if str(ws[f'P{i}'].value)[0:4] == 'АОСР' or str(ws[f'R{i}'].value)[0:4]   == 'АОСР':
                names_aosr.add(ws[f'C{i}'].value)
            elif str(ws[f'P{i}'].value)[0:4] == 'АРОО' or str(ws[f'R{i}'].value)[0:4] == 'АРОО':
                names_aroo.add(ws[f'C{i}'].value)
            elif str(ws[f'P{i}'].value)[0:4] == 'АООК' or str(ws[f'R{i}'].value)[0:4] == 'АООК':
                names_aook.add(ws[f'C{i}'].value)
    return names_aosr, names_aook, names_aroo
def change_names_file(names_aosr, names_aook, names_aroo):
    xls = load_workbook(r'names\names.xlsx')
    ws = xls.active
    ws['A1'].value = 'Список названий АОСР'
    ws['C1'].value = 'Список названий АООК'
    ws['E1'].value = 'Список названий АРОО'
    for counter, name in enumerate(names_aosr):
        ws[f'A{counter+2}'].value = name
    for counter, name in enumerate(names_aook):
        ws[f'C{counter+2}'].value = name
    for counter, name in enumerate(names_aroo):
        ws[f'D{counter+2}'].value = name
    xls.save(r'names\names.xlsx')

if __name__ == '__main__':
    names_aosr, names_aook, names_aroo = create_names_from_excel_files()
    change_names_file(names_aosr = names_aosr, names_aook = names_aook, names_aroo = names_aroo)

