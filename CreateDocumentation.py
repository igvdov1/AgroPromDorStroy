from openpyxl import load_workbook, Workbook
import os
from GetNamesFromExcel import GetNamesFromExcel
from CreateDocument import CreateDocument

class CreateFullDocumentation():
    def __init__(self, links_to_smeta, link):
        self.links_to_smeta = links_to_smeta
        self.names_aosr, self.names_aook, self.names_aroo = GetNamesFromExcel().load_names()
        self.link = link
        self.counter_aosr = 1
        self.counter_aook = 1
        self.counter_aroo = 1
    def create_docs(self):
        for link in self.links_to_smeta:
            self.xls = load_workbook(link)
            self.wb = self.xls.active
            for i in range(1, self.wb.max_row):
                if self.wb[f'C{i}'].value in self.names_aosr:
                    CreateDocument(name = self.wb[f'C{i}'].value, type = 'АОСР', link = self.link, number=self.counter_aosr).create_document()
                    self.counter_aosr += 1
                elif self.wb[f'C{i}'].value in self.names_aook:
                    CreateDocument(name = self.wb[f'C{i}'].value, type = 'АООК', link = self.link, number=self.counter_aook).create_document()
                    self.counter_aook += 1
                elif self.wb[f'C{i}'].value in self.names_aroo:
                    CreateDocument(name = self.wb[f'C{i}'], type = 'АРОО', link = self.link, number = self.counter_aroo).create_document()
                    self.counter_aroo += 1
                




