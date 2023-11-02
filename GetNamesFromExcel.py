import os
from openpyxl import load_workbook
path = r'C:\Users\igvdo\.vscode\app_for_agroprom\names\names.xlsx'

class GetNamesFromExcel():
    def __init__(self):
        self.xls = load_workbook(path)
        self.ws = self.xls.active
    def load_names(self):
        letters = ['A', 'C', 'E']
        names_aosr = set()
        names_aook = set()
        names_aroo = set()
        for letter in letters:
            for i in range(2, len(self.ws[f'{letter}'])+1):
                if letter == 'A':
                    names_aosr.add(self.ws[f'{letter}{i}'].value)
                if letter == 'C':
                    names_aook.add(self.ws[f'{letter}{i}'].value)
                if letter == 'E':
                    names_aroo.add(self.ws[f'{letter}{i}'].value)
        names_aosr.discard(None)
        names_aook.discard(None)
        names_aroo.discard(None)
        return names_aosr, names_aook, names_aroo
                    
