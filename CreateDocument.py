from openpyxl import load_workbook
import os

class CreateDocument():
    def __init__(self, name, type, link, number, additional_materials):
        self.type = type
        self.name = name
        self.number = number
        self.template = load_workbook('template_excel.xlsx')
        self.ws = self.template.active
        self.link = link
        self.additional_materials = additional_materials
    def create_document(self):
        for i in range(1, self.ws.max_row):
            if self.ws[f'B{i}'].value == 'К освидетельствованию предъявлены следующие работы':
                self.ws[f'I{i}'].value = self.name
                
            if self.ws[f'A{i}'].value == 'Объект капитального строительства ':
                self.ws[f'F{i}'].value = self.additional_materials['Объект Капитального'].text()
            if self.ws[f'A{i}'].value == 'Застройщик (технический заказчик, эксплуатирующая организация или региональный оператор)':
                self.ws[f'A{i+1}'].value = self.additional_materials['Застройщик'].text()
            if self.ws[f'A{i}'].value == 'Лицо, осуществляющее строительство':
                self.ws[f'F{i}'].value = self.additional_materials['Лицо осущ строй'].text()
            if self.ws[f'A{i}'].value == 'Лицо, осуществляющее подготовку проектной документации ':
                self.ws[f'K{i}'].value = self.additional_materials['Лицо осущ документ'].text()
            if self.ws[f'A{i}'].value == '№':
                self.ws[f'B{i}'].value = self.number
            if self.ws[f'A{i}'].value == 'Представитель застройщика (технического заказчика, эксплуатирующей организации или регионального оператора) по вопросам строительного контроля ':
                self.ws[f'A{i+1}'].value = self.additional_materials['Представитель застройщика'].text()    
            if self.ws[f'A{i}'].value == 'Представитель лица, осуществляющего строительство':
                self.ws[f'A{i+1}'].value = self.additional_materials['Представитель лица строительства'].text() 
            if self.ws[f'A{i}'].value == 'Представитель лица, осуществляющего строительство, по вопросам строительного контроля (специалист по организации строительства)':
                self.ws[f'A{i+1}'].value = self.additional_materials['Представитель лица строительства2'].text()  
            if self.ws[f'A{i}'].value == 'Представитель лица, осуществляющего подготовку проектной документации':
                self.ws[f'A{i+1}'].value = self.additional_materials['Представитель лица документации'].text() 
            if self.ws[f'A{i}'].value == 'Представитель лица, выполнившего работы, подлежащие освидетельствованию ':
                self.ws[f'A{i+1}'].value = self.additional_materials['Представитель лица работы'].text()           
            if self.ws[f'A{i}'].value == 'а также иные представители лиц, участвующих в освидетельствовании':
                self.ws[f'A{i+1}'].value = self.additional_materials['иные представители'].text()
            if self.ws[f'B{i}'].value == 'Работы выполнены по проектной документации':
                self.ws[f'A{i+1}'].value = self.additional_materials['Работы выполнены пд'].text()
            if self.ws[f'B{i}'].value == 'При выполнении работ применены:':
                self.ws[f'A{i+1}'].value = self.additional_materials['При выполнении применены'].text()
            if self.ws[f'B{i}'].value == 'Предъявлены документы, подтверждающие соответствие работ предъявляемым к ним требованиям:    ':
                self.ws[f'A{i+1}'].value = self.additional_materials['Предъявлены доки'].text()
            if self.ws[f'B{i}'].value == 'Работы выполнены в соответствии с':
                self.ws[f'A{i+1}'].value = self.additional_materials['Работы выполнены пд'].text() + ' СП ...'
            if self.ws[f'B{i}'].value == 'Разрешается производство последующих работ:':
            #Вот тут добавить название следующей АОСР
                self.ws[f'H{i}'].value = self.name
            if self.ws[f'A{i}'].value == 'Акт составлен в ':
                self.ws[f'D{i}'].value = '2 (двух)'
            if self.ws[f'A{i}'].value == 'Приложения:':
                self.ws[f'A{i+1}'].value = self.additional_materials['приложения'].text()
        
        
        
        os.mkdir(fr'{self.link}\{self.type} №{self.number}')
        self.template.save(fr'{self.link}\{self.type} №{self.number}\{self.type} {self.number}test.xlsx')
